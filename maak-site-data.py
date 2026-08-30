# -*- coding: utf-8 -*-
"""Leest data/podcast-liveshows.xlsx en schrijft data/site-data.js.

Draai dit script elke keer nadat je iets in het Excel-bestand hebt aangepast:

    python3 maak-site-data.py

De site leest site-data.js, niet het Excel-bestand zelf. Dat is nodig omdat een
webpagina die je met dubbelklikken opent geen los bestand van je schijf mag
inlezen (een beveiliging van de browser).
"""
import json, os, sys
from openpyxl import load_workbook

HIER = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HIER, "data", "podcast-liveshows.xlsx")
UIT  = os.path.join(HIER, "data", "site-data.js")

def lees(wb, naam):
    ws = wb[naam]
    kop = [c.value for c in ws[1]]
    rijen = []
    for r in ws.iter_rows(min_row=2):
        w = [c.value for c in r]
        if all(v is None or str(v).strip() == "" for v in w):
            continue
        rijen.append({k: (v if v is not None else "") for k, v in zip(kop, w)})
    return rijen

if not os.path.exists(XLSX):
    sys.exit("Kan %s niet vinden." % XLSX)

wb = load_workbook(XLSX, data_only=True)
data = {naam: lees(wb, naam) for naam in
        ["podcasts", "shows", "show_podcasts", "events", "venues"]}

with open(UIT, "w", encoding="utf-8") as f:
    f.write("// Automatisch gemaakt door maak-site-data.py - niet met de hand aanpassen.\n")
    f.write("window.DATA = ")
    json.dump(data, f, ensure_ascii=False, indent=1, default=str)
    f.write(";\n")

met_coord = sum(1 for v in data["venues"]
                if str(v.get("lat", "")).strip() and str(v.get("lon", "")).strip())
print("Weggeschreven naar data/site-data.js")
for naam in data:
    print("  %-14s %3d rijen" % (naam, len(data[naam])))
print("  zalen met lat/lon: %d van %d" % (met_coord, len(data["venues"])))
