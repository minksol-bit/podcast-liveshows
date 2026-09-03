# -*- coding: utf-8 -*-
"""Bouwt de hele website uit data/podcast-liveshows.xlsx.

Draai dit na elke wijziging in het Excel-bestand:

    python3 bouw-site.py

Wat er gemaakt wordt:
    index.html          de agenda met kaart
    catalogus.html      alle podcasts als blokken
    toplijst.html       de Apple top 100 met wie er live speelt
    podcast/<naam>.html een pagina per podcast
    data/site-data.js   de data die de agenda inleest
    sitemap.xml         lijst van pagina's voor Google

Pas de pagina's niet met de hand aan: dit script overschrijft ze.
Vormgeving zit in assets/stijl.css, dat blijft ongemoeid.
"""
import json, os, re, sys, unicodedata, html
from datetime import date, datetime
from openpyxl import load_workbook

HIER = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HIER, "data", "podcast-liveshows.xlsx")
TOP100 = os.path.join(HIER, "data", "apple-top100.json")

# Zet hier het echte webadres zodra de site online staat.
SITE_URL = "https://podcast-liveshows.nl"

MAANDEN = ["januari","februari","maart","april","mei","juni",
           "juli","augustus","september","oktober","november","december"]
MAAND_KORT = ["jan","feb","mrt","apr","mei","jun","jul","aug","sep","okt","nov","dec"]

# ---------------------------------------------------------------- hulpjes

def e(s):
    return html.escape("" if s is None else str(s), quote=True)

def helderheid(hexkleur):
    # Relatieve helderheid van een hex-kleur (0 = zwart, 1 = wit), om te bepalen
    # of witte of donkere tekst leesbaar is op die achtergrond.
    h = (hexkleur or "").lstrip("#")
    if len(h) != 6:
        return 0.3
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b) / 255

def slug(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "podcast"

def zoeksleutel(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    s = s.replace("&", " en ")
    s = re.sub(r"\b(de|het|een|the|en|and)\b", " ", s)
    return re.sub(r"[^a-z0-9]+", "", s)

def leeg(v):
    return v is None or str(v).strip() == ""

def tekst(v):
    return "" if leeg(v) else str(v).strip()

def getal(v):
    if leeg(v): return None
    try: return float(str(v).replace(",", "."))
    except ValueError: return None

def ontleed_datum(waarde):
    if isinstance(waarde, datetime):
        return {"jaar": waarde.year, "maand": waarde.month, "dag": waarde.day,
                "tijd": waarde.strftime("%H:%M") if (waarde.hour or waarde.minute) else None}
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})(?:[ T](\d{2}):(\d{2}))?", str(waarde).strip())
    if not m: return None
    return {"jaar": int(m.group(1)), "maand": int(m.group(2)), "dag": int(m.group(3)),
            "tijd": (m.group(4) + ":" + m.group(5)) if m.group(4) else None}

def iso(d):
    return "%04d-%02d-%02d" % (d["jaar"], d["maand"], d["dag"])

def datum_lang(d):
    return "%d %s %d" % (d["dag"], MAANDEN[d["maand"] - 1], d["jaar"])

def datum_kort(d):
    return "%d %s" % (d["dag"], MAAND_KORT[d["maand"] - 1])

# ---------------------------------------------------------------- inlezen

def tabblad(wb, naam):
    ws = wb[naam]
    kop = [c.value for c in ws[1]]
    uit = []
    for r in ws.iter_rows(min_row=2):
        w = [c.value for c in r]
        if all(leeg(v) for v in w): continue
        uit.append(dict(zip(kop, w)))
    return uit

def lees_alles():
    if not os.path.exists(XLSX):
        sys.exit("Kan %s niet vinden." % XLSX)
    wb = load_workbook(XLSX, data_only=True)
    rauw = {n: tabblad(wb, n) for n in
            ["podcasts", "shows", "show_podcasts", "events", "venues"]}
    rauw["toplijst"] = tabblad(wb, "toplijst") if "toplijst" in wb.sheetnames else []

    podcasts = {}
    for p in rauw["podcasts"]:
        podcasts[p["id"]] = {
            "id": p["id"], "naam": tekst(p.get("naam")),
            "cover": tekst(p.get("cover_url")), "thema": tekst(p.get("thema")) or "Overig",
            "kort": tekst(p.get("omschrijving")), "lang": tekst(p.get("omschrijving_lang")),
            "apple_id": tekst(p.get("apple_id")), "rang": p.get("apple_rang"),
            "website": tekst(p.get("website")),
            "band_links": tekst(p.get("bannerkleur_links")) or "#8e212e",
            "band_rechts": tekst(p.get("bannerkleur_rechts")) or "#6d1823",
            "slug": slug(p.get("naam")), "shows": [], "events": [],
        }

    shows = {}
    for s in rauw["shows"]:
        shows[s["id"]] = {"id": s["id"], "titel": tekst(s.get("titel")),
                          "type": tekst(s.get("type")), "organisator": tekst(s.get("organisator")),
                          "podcasts": []}

    for k in rauw["show_podcasts"]:
        s, p = shows.get(k.get("show_id")), podcasts.get(k.get("podcast_id"))
        if s and p:
            s["podcasts"].append(p)
            if s not in p["shows"]: p["shows"].append(s)

    venues = {}
    for v in rauw["venues"]:
        la, lo = getal(v.get("lat")), getal(v.get("lon"))
        venues[v["id"]] = {"id": v["id"], "naam": tekst(v.get("naam")), "stad": tekst(v.get("stad")),
                           "provincie": tekst(v.get("provincie")), "lat": la, "lon": lo,
                           "opkaart": la is not None and lo is not None}

    events = []
    for ev in rauw["events"]:
        d = ontleed_datum(ev.get("aanvang"))
        show, zaal = shows.get(ev.get("show_id")), venues.get(ev.get("venue_id"))
        if not d or not show or not zaal: continue
        rij = {"id": ev["id"], "d": d, "show": show, "zaal": zaal,
               "iso": iso(d), "tijd": d["tijd"],
               "ticket": tekst(ev.get("ticket_url")), "bron": tekst(ev.get("bron_url")),
               "status": tekst(ev.get("status")), "prijs": getal(ev.get("prijs_vanaf")),
               "gecheckt": tekst(ev.get("laatst_gecheckt")),
               "maand": "%04d-%02d" % (d["jaar"], d["maand"]),
               "provincie": zaal["provincie"],
               "themas": sorted({p["thema"] for p in show["podcasts"]})}
        events.append(rij)
        for p in show["podcasts"]:
            p["events"].append(rij)
    events.sort(key=lambda x: (x["iso"], x["tijd"] or "00:00"))
    for p in podcasts.values():
        p["events"].sort(key=lambda x: (x["iso"], x["tijd"] or "00:00"))

    status = {}
    for r in rauw["toplijst"]:
        status[str(r.get("apple_id") or "")] = {
            "gecontroleerd": tekst(r.get("gecontroleerd_op")),
            "bevinding": tekst(r.get("bevinding")),
            "notitie": tekst(r.get("notitie"))}
    return podcasts, shows, venues, events, status

# ---------------------------------------------------------------- sjablonen

def render(sjabloon, **kv):
    for k, v in kv.items():
        sjabloon = sjabloon.replace("{{%s}}" % k, str(v))
    return sjabloon

KOP = """<!DOCTYPE html>
<html lang="nl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{{titel}}</title>
<meta name="description" content="{{omschrijving}}">
<link rel="canonical" href="{{canoniek}}">
<meta property="og:title" content="{{titel}}">
<meta property="og:description" content="{{omschrijving}}">
<meta property="og:type" content="website">
<link rel="icon" type="image/png" href="{{basis}}assets/favicon.png">
<link rel="stylesheet" href="{{basis}}assets/stijl.css">
{{extra_head}}
</head>
<body>

<div class="balk">
  <div class="wrap">
    <a class="merk" href="{{basis}}index.html">
      <img src="{{basis}}assets/logo-mark.png" alt="Podcast Liveshows">
      <span>Podcast Liveshows</span>
    </a>
    <nav class="menu">
      <a href="{{basis}}index.html" class="{{m_agenda}}">Agenda</a>
      <a href="{{basis}}catalogus.html" class="{{m_catalogus}}">Podcasts</a>
      <a href="{{basis}}toplijst.html" class="{{m_toplijst}}">Toplijst</a>
    </nav>
  </div>
</div>

<div class="wrap">
"""

VOET = """
  <footer>
    Kaartgegevens van <a href="https://www.openstreetmap.org/copyright" rel="noopener">OpenStreetMap</a>.
    Ranglijst van Apple Podcasts. Agenda voor het laatst gecontroleerd op {{gecheckt}}.
  </footer>
</div>
{{scripts}}
</body>
</html>
"""

def kop(titel, omschrijving, basis="", actief="", extra_head="", pad=""):
    return render(KOP, titel=e(titel), omschrijving=e(omschrijving), basis=basis,
                  canoniek=e(SITE_URL.rstrip("/") + "/" + pad),
                  extra_head=extra_head,
                  m_agenda="hier" if actief == "agenda" else "",
                  m_catalogus="hier" if actief == "catalogus" else "",
                  m_toplijst="hier" if actief == "toplijst" else "")

def voet(gecheckt, scripts=""):
    return render(VOET, gecheckt=e(gecheckt or "-"), scripts=scripts)

def leaflet_head(basis=""):
    return ('<link rel="stylesheet" href="%sassets/leaflet/leaflet.css">\n'
            '<link rel="stylesheet" href="%sassets/leaflet/MarkerCluster.css">' % (basis, basis))

def leaflet_scripts(basis=""):
    return ('<script src="%sassets/leaflet/leaflet.js"></script>\n'
            '<script src="%sassets/leaflet/leaflet.markercluster.js"></script>' % (basis, basis))

HERO = """
  <div class="hero{{klein}}">
    <div class="hero-beeld" id="hero-beeld"></div>
    <div class="hero-glow" id="hero-glow"></div>
    <div class="ring r1"></div>
    <div class="ring r2"></div>
    <div class="hero-fade"></div>
    <div class="hero-tekst">
      <h1>{{kopregel}}</h1>
      {{onder}}
    </div>
  </div>
"""

def hero(kopregel, onder="", klein=False):
    return render(HERO, kopregel=kopregel, onder=onder, klein=" klein" if klein else "")

PARALLAX = """
<script>
(function () {
  if (window.matchMedia && window.matchMedia("(prefers-reduced-motion: reduce)").matches) return;
  var beeld = document.getElementById("hero-beeld"), glow = document.getElementById("hero-glow"), bezig = false;
  if (!beeld) return;
  window.addEventListener("scroll", function () {
    if (bezig) return;
    bezig = true;
    window.requestAnimationFrame(function () {
      var y = window.pageYOffset || 0;
      if (y < 900) {
        beeld.style.transform = "translate3d(0," + (y * 0.22) + "px,0)";
        if (glow) glow.style.transform = "translate3d(0," + (y * 0.10) + "px,0)";
      }
      bezig = false;
    });
  }, { passive: true });
})();
</script>
"""

def jsonld_event(ev, url):
    d = {"@context": "https://schema.org", "@type": "Event",
         "name": ev["show"]["titel"],
         "startDate": ev["iso"] + ("T" + ev["tijd"] + ":00" if ev["tijd"] else ""),
         "eventAttendanceMode": "https://schema.org/OfflineEventAttendanceMode",
         "eventStatus": "https://schema.org/EventScheduled",
         "url": url,
         "location": {"@type": "Place", "name": ev["zaal"]["naam"],
                      "address": {"@type": "PostalAddress",
                                  "addressLocality": ev["zaal"]["stad"],
                                  "addressRegion": ev["zaal"]["provincie"],
                                  "addressCountry": "NL"}}}
    if ev["zaal"]["opkaart"]:
        d["location"]["geo"] = {"@type": "GeoCoordinates",
                                "latitude": ev["zaal"]["lat"], "longitude": ev["zaal"]["lon"]}
    if ev["show"]["podcasts"]:
        d["performer"] = [{"@type": "PerformingGroup", "name": p["naam"]} for p in ev["show"]["podcasts"]]
    if ev["ticket"]:
        aanbod = {"@type": "Offer", "url": ev["ticket"], "availability": "https://schema.org/InStock"}
        if ev["prijs"] is not None:
            aanbod["price"] = "%.2f" % ev["prijs"]
            aanbod["priceCurrency"] = "EUR"
        d["offers"] = aanbod
    return d

# ---------------------------------------------------------------- pagina's

def lees_top100():
    if not os.path.exists(TOP100): return None
    return json.load(open(TOP100, encoding="utf-8"))

def cover_van(p, maat="600x600"):
    return p["cover"] or ""

def schrijf(pad, inhoud):
    vol = os.path.join(HIER, pad)
    os.makedirs(os.path.dirname(vol), exist_ok=True)
    with open(vol, "w", encoding="utf-8") as f:
        f.write(inhoud)
    return vol

def data_voor_agenda(podcasts, venues, events):
    rijen = []
    for ev in events:
        rijen.append({
            "id": ev["id"], "iso": ev["iso"], "dag": ev["d"]["dag"],
            "maandnr": ev["d"]["maand"], "jaar": ev["d"]["jaar"], "maand": ev["maand"],
            "tijd": ev["tijd"], "titel": ev["show"]["titel"],
            "prijs": ev["prijs"], "ticket": ev["ticket"], "status": ev["status"],
            "provincie": ev["provincie"], "themas": ev["themas"],
            "zaal": {"id": ev["zaal"]["id"], "naam": ev["zaal"]["naam"], "stad": ev["zaal"]["stad"],
                     "opkaart": ev["zaal"]["opkaart"], "lat": ev["zaal"]["lat"], "lon": ev["zaal"]["lon"]},
            "podcasts": [{"naam": p["naam"], "slug": p["slug"], "cover": p["cover"]}
                         for p in ev["show"]["podcasts"]],
        })
    zonder = sum(1 for v in venues.values() if not v["opkaart"])
    return {"events": rijen, "venues_totaal": len(venues), "venues_zonder_coordinaten": zonder}

FILTERBLOK = """
  <div class="filters">
    <div class="filter filter-veelkeuze" id="fm-maand">
      <label id="fm-maand-label">Maand</label>
      <button type="button" class="veelkeuze-knop" id="fm-maand-knop" aria-haspopup="true"
        aria-expanded="false" aria-labelledby="fm-maand-label fm-maand-knop">Alle maanden</button>
      <div class="veelkeuze-paneel" id="fm-maand-paneel" hidden></div>
    </div>
    <div class="filter filter-veelkeuze" id="fm-provincie">
      <label id="fm-provincie-label">Provincie</label>
      <button type="button" class="veelkeuze-knop" id="fm-provincie-knop" aria-haspopup="true"
        aria-expanded="false" aria-labelledby="fm-provincie-label fm-provincie-knop">Alle provincies</button>
      <div class="veelkeuze-paneel" id="fm-provincie-paneel" hidden></div>
    </div>
    <div class="filter filter-veelkeuze" id="fm-thema">
      <label id="fm-thema-label">Thema</label>
      <button type="button" class="veelkeuze-knop" id="fm-thema-knop" aria-haspopup="true"
        aria-expanded="false" aria-labelledby="fm-thema-label fm-thema-knop">Alle thema's</button>
      <div class="veelkeuze-paneel" id="fm-thema-paneel" hidden></div>
    </div>
    <div class="filter"><label for="f-prijs">Prijs</label>
      <select id="f-prijs">
        <option value="">Alle prijzen</option>
        <option value="20">tot &euro;20</option>
        <option value="30">tot &euro;30</option>
        <option value="40">tot &euro;40</option>
        <option value="50">tot &euro;50</option>
      </select></div>
    <div class="filter"><label>Snel</label>
      <div class="snel">
        <button type="button" id="snel-vandaag">Vanavond</button>
        <button type="button" id="snel-weekend">Dit weekend</button>
        <button type="button" id="snel-fav">Favorieten</button>
      </div></div>
    <div class="filter"><label>&nbsp;</label><button type="button" id="wis">Wis filters</button></div>
    <div class="telling" id="telling"></div>
  </div>
"""

def bouw_index(podcasts, venues, events, gecheckt):
    data = data_voor_agenda(podcasts, venues, events)
    schrijf("data/site-data.js",
            "// Automatisch gemaakt door bouw-site.py - niet met de hand aanpassen.\n"
            "window.DATA = " + json.dumps(data, ensure_ascii=False, indent=1, default=str) + ";\n")

    komend = [ev for ev in events][:60]
    ld = json.dumps([jsonld_event(ev, SITE_URL.rstrip("/") + "/index.html") for ev in komend],
                    ensure_ascii=False, indent=1)
    extra = '<script type="application/ld+json">%s</script>' % ld

    onder = ('<p class="intro">Steeds meer podcasts stappen het theater in. Deze site verzamelt welke '
             'Nederlandse podcasts een liveshow spelen, wanneer en in welke zaal, met een directe link '
             'naar de kaartverkoop. Filter op maand, provincie of thema om te zien wat er bij jou in de '
             'buurt te doen is.</p>'
             '<p class="cijfers">%d liveshows van %d podcasts in %d zalen.</p>'
             % (len(events), len(podcasts), len(venues)))

    html_uit = (kop("Podcast Liveshows in Nederland",
                    "Welke Nederlandse podcasts spelen een liveshow, wanneer en waar. "
                    "Agenda en kaart met directe link naar de kaartverkoop.",
                    basis="", actief="agenda", extra_head=leaflet_head() + "\n" + extra,
                    pad="index.html")
                + hero("Welke podcast staat er<br><span class=\"accent\">bij jou in het theater?</span>", onder)
                + FILTERBLOK
                + """
  <div class="kolommen">
    <div class="lijstkolom"><div id="lijst"></div></div>
    <div class="kaartkolom">
      <div class="kaart" id="kaart"></div>
      <div class="melding" id="melding" hidden></div>
    </div>
  </div>
"""
                + voet(gecheckt,
                       leaflet_scripts()
                       + '\n<script src="assets/favorieten.js"></script>'
                       + '\n<script src="assets/veelkeuze.js"></script>'
                       + '\n<script src="data/site-data.js"></script>'
                       + '\n<script src="assets/agenda.js"></script>'
                       + PARALLAX))
    schrijf("index.html", html_uit)

def bouw_catalogus(podcasts, gecheckt):
    lijst = sorted(podcasts.values(), key=lambda p: p["naam"].lower())
    themas = sorted({p["thema"] for p in lijst})
    stukken = []
    huidige_letter = None
    for p in lijst:
        letter = p["naam"][0].upper()
        if not letter.isalpha(): letter = "#"
        if letter != huidige_letter:
            huidige_letter = letter
            stukken.append('    <div class="letterkop">%s</div>' % e(letter))
        aantal = len(p["events"])
        regel = ("%d liveshow%s" % (aantal, "" if aantal == 1 else "s")) if aantal else "nog geen liveshow bekend"
        beeld = ('<img src="%s" alt="" loading="lazy">' % e(p["cover"])) if p["cover"] else ""
        stukken.append(
            '    <a class="blok" href="podcast/%s.html" data-naam="%s" data-maker="%s" data-thema="%s">\n'
            '      %s\n'
            '      <div class="naamplaat">%s</div>\n'
            '      <div class="over">\n'
            '        <div class="merkje">%s</div>\n'
            '        <div class="kop">%s</div>\n'
            '        <div class="tekst">%s</div>\n'
            '        <div class="voet">%s &rarr;</div>\n'
            '      </div>\n'
            '    </a>'
            % (e(p["slug"]), e(p["naam"]), e(p.get("maker", "")), e(p["thema"]),
               beeld, e(p["naam"]), e(p["thema"]), e(p["naam"]),
               e(p["kort"] or "Nog geen omschrijving."), e(regel)))

    filters = ('  <div class="filters">\n'
               '    <div class="filter"><label for="zoek">Zoeken</label>'
               '<input type="search" id="zoek" placeholder="Naam of maker"></div>\n'
               '    <div class="filter filter-veelkeuze" id="fm-thema">\n'
               '      <label id="fm-thema-label">Thema</label>\n'
               '      <button type="button" class="veelkeuze-knop" id="fm-thema-knop" aria-haspopup="true" '
               'aria-expanded="false" aria-labelledby="fm-thema-label fm-thema-knop">Alle thema\'s</button>\n'
               '      <div class="veelkeuze-paneel" id="fm-thema-paneel" hidden>%s</div>\n'
               '    </div>\n'
               '    <div class="telling" id="telling"></div>\n'
               '  </div>\n'
               % "".join('<label class="veelkeuze-optie"><input type="checkbox" value="%s"> %s</label>'
                         % (e(t), e(t)) for t in themas))

    html_uit = (kop("Alle podcasts - Podcast Liveshows",
                    "Overzicht van alle podcasts op deze site, van A tot Z, met hun liveshows.",
                    basis="", actief="catalogus", pad="catalogus.html")
                + hero("Alle podcasts", '<p class="intro">Van A tot Z. Zweef over een blok voor een korte '
                       'omschrijving, klik erop voor alle shows en de kaart.</p>', klein=True)
                + filters
                + '  <div class="raster">\n' + "\n".join(stukken) + "\n  </div>\n"
                + voet(gecheckt, '<script src="assets/veelkeuze.js"></script>\n<script src="assets/catalogus.js"></script>' + PARALLAX))
    schrijf("catalogus.html", html_uit)

def bouw_toplijst(podcasts, status, gecheckt):
    top = lees_top100()
    if not top:
        print("  (geen apple-top100.json gevonden, toplijst overgeslagen)")
        return 0, 0
    op_sleutel = {zoeksleutel(p["naam"]): p for p in podcasts.values()}
    rijen, met_live, nagekeken = [], 0, 0
    for rang, naam, maker, genres, apple_id, beeld in top["lijst"]:
        eigen = op_sleutel.get(zoeksleutel(naam))
        heeft = eigen and eigen["events"]
        st = status.get(str(apple_id), {})
        if st.get("gecontroleerd"):
            nagekeken += 1
        beeld_url = top["beeld_basis"] + beeld + "/200x200bb.png"
        binnen = ('<div class="rang">%d</div>'
                  '<img src="%s" alt="" loading="lazy">'
                  '<div class="mid"><div class="naam">%s</div><div class="maker">%s</div></div>'
                  % (rang, e(beeld_url), e(naam), e(maker)))
        if heeft:
            met_live += 1
            staat = "%d liveshow%s" % (len(eigen["events"]), "" if len(eigen["events"]) == 1 else "s")
            rijen.append('    <a class="toprij live" href="podcast/%s.html">%s<div class="staat">%s &rarr;</div></a>'
                         % (e(eigen["slug"]), binnen, e(staat)))
        elif st.get("gecontroleerd"):
            rijen.append('    <div class="toprij stil" title="%s">%s<div class="staat">nagekeken %s '
                         '&middot; geen liveshow</div></div>'
                         % (e(st.get("notitie", "")), binnen, e(st["gecontroleerd"])))
        else:
            rijen.append('    <div class="toprij onbekend">%s<div class="staat">nog niet bekeken</div></div>'
                         % binnen)

    onder = ('<p class="intro">De honderd best beluisterde podcasts van Nederland volgens Apple Podcasts, '
             'bijgewerkt op %s. Wie een liveshow in onze agenda heeft is aanklikbaar. De rest is of '
             'nagekeken zonder dat we een voorstelling vonden, of staat nog op de lijst om uit te zoeken.</p>'
             '<p class="cijfers">%d van de 100 nagekeken &middot; %d met een liveshow &middot; '
             '%d nog te doen</p>'
             % (e(top.get("opgehaald", "")), nagekeken, met_live, 100 - nagekeken))

    html_uit = (kop("Toplijst - Podcast Liveshows",
                    "De top 100 podcasts van Nederland volgens Apple Podcasts, met wie er live in het theater staat.",
                    basis="", actief="toplijst", pad="toplijst.html")
                + hero("De top 100, en wie er<br><span class=\"accent\">live te zien is</span>", onder, klein=True)
                + '  <div class="toplijst">\n' + "\n".join(rijen) + "\n  </div>\n"
                + voet(gecheckt, PARALLAX))
    schrijf("toplijst.html", html_uit)
    return met_live, nagekeken

def bouw_podcastpaginas(podcasts, gecheckt):
    # Verweesde pagina's opruimen: een podcast die samengevoegd of verwijderd is
    # (zie CLAUDE.md, "live" in de matchsleutel) laat anders een oude pagina achter
    # die nergens meer naar linkt maar wel online blijft staan.
    huidige_slugs = {p["slug"] for p in podcasts.values()}
    podcastmap = os.path.join(HIER, "podcast")
    if os.path.isdir(podcastmap):
        for bestand in os.listdir(podcastmap):
            if bestand.endswith(".html") and bestand[:-5] not in huidige_slugs:
                os.remove(os.path.join(podcastmap, bestand))
    gemaakt = 0
    for p in podcasts.values():
        evs = p["events"]
        # Een zaal kan meerdere keren voorkomen. Speldjes op exact dezelfde plek
        # vallen over elkaar heen, dus we zetten er een per zaal neer en noemen
        # alle datums in de ballon.
        zalen = {}
        for ev in evs:
            if not ev["zaal"]["opkaart"]: continue
            z = zalen.setdefault(ev["zaal"]["id"], {
                "id": ev["zaal"]["id"], "naam": ev["zaal"]["naam"], "stad": ev["zaal"]["stad"],
                "lat": ev["zaal"]["lat"], "lon": ev["zaal"]["lon"], "datums": []})
            z["datums"].append({"ev": ev["id"], "dag": ev["d"]["dag"], "maandnr": ev["d"]["maand"],
                                "jaar": ev["d"]["jaar"], "tijd": ev["tijd"]})
        data = {"naam": p["naam"], "zalen": list(zalen.values()), "events": [
            {"id": ev["id"], "iso": ev["iso"], "dag": ev["d"]["dag"], "maandnr": ev["d"]["maand"],
             "jaar": ev["d"]["jaar"], "tijd": ev["tijd"], "zaal_id": ev["zaal"]["id"],
             "zaal": {"naam": ev["zaal"]["naam"], "stad": ev["zaal"]["stad"],
                      "opkaart": ev["zaal"]["opkaart"], "lat": ev["zaal"]["lat"], "lon": ev["zaal"]["lon"]}}
            for ev in evs]}

        rijen = []
        vorige = None
        for ev in evs:
            if ev["maand"] != vorige:
                vorige = ev["maand"]
                rijen.append('    <h2 class="maand">%s %d</h2>' % (MAANDEN[ev["d"]["maand"] - 1], ev["d"]["jaar"]))
            json_ev = json.dumps({"id": ev["id"], "iso": ev["iso"], "tijd": ev["tijd"],
                                  "titel": ev["show"]["titel"], "zaal": ev["zaal"]["naam"],
                                  "stad": ev["zaal"]["stad"], "ticket": ev["ticket"]}, ensure_ascii=False)
            prijstekst = ('vanaf &euro;%s'
                     % ("%.2f" % ev["prijs"]).replace(".", ",").replace(",00", ",-")) if ev["prijs"] is not None else ""
            onder = ('<span class="knop-onder">%s</span>' % prijstekst) if prijstekst else ""
            if ev["status"].lower() == "uitverkocht":
                knop = '<span class="geen knop-vorm"><span class="knop-label">Uitverkocht</span>%s</span>' % onder
            elif ev["ticket"]:
                knop = ('<a class="knop" href="%s" target="_blank" rel="noopener">'
                        '<span class="knop-label">Tickets</span>%s</a>' % (e(ev["ticket"]), onder))
            else:
                knop = '<span class="geen knop-vorm"><span class="knop-label">geen link</span>%s</span>' % onder
            rijen.append(
                '    <div class="event" data-ev="%s" data-json="%s">\n'
                '      <div class="datum"><div class="dag">%d</div><div class="mnd">%s</div></div>\n'
                '      <div class="info"><div class="titel">%s</div>'
                '<div class="zaal">%s, %s%s</div>'
                '<div class="bij">%s</div></div>\n'
                '      <div class="rechtsblok">%s</div>\n'
                '    </div>'
                % (e(ev["id"]), e(json_ev), ev["d"]["dag"], MAAND_KORT[ev["d"]["maand"] - 1],
                   e(ev["show"]["titel"]), e(ev["zaal"]["naam"]), e(ev["zaal"]["stad"]),
                   (' <span style="opacity:.7">(%s)</span>' % e(ev["provincie"])) if ev["provincie"] else "",
                   ("Aanvang " + ev["tijd"]) if ev["tijd"] else "&nbsp;",
                   knop))

        heeft_kaart = any(ev["zaal"]["opkaart"] for ev in evs)
        n_zalen = len({ev["zaal"]["id"] for ev in evs if ev["zaal"]["opkaart"]})
        kaartblok = ('    <div class="kaartkolom"><div class="kaart" id="kaart"></div>'
                     '<p class="cijfers" style="margin-top:10px">%d show%s in %d zaal%s. Speelt een podcast '
                     'twee keer in dezelfde zaal, dan staan beide datums in hetzelfde speldje.</p></div>\n'
                     % (len(evs), "" if len(evs) == 1 else "s", n_zalen, "" if n_zalen == 1 else "en")
                     if heeft_kaart else
                     '    <div class="kaartkolom"><div class="melding">Nog geen coördinaten voor de zalen '
                     'van deze podcast, dus nog geen kaart.</div></div>\n')

        band_stijl = "background:linear-gradient(90deg,%s,%s)" % (e(p["band_links"]), e(p["band_rechts"]))
        pbanner_klasse = "pbanner pbanner-licht" if helderheid(p["band_rechts"]) > 0.6 else "pbanner pbanner-donker"
        if p["cover"]:
            logo = '<img class="pbanner-logo" src="%s" alt="">' % e(p["cover"])
        else:
            logo = '<div class="pbanner-logo pbanner-logo-leeg"></div>'
        rang = ('<span class="merkje">#%s in de Apple top 100</span>' % e(p["rang"])) if p["rang"] else ""
        maker = ('<p class="maker">%s</p>' % e(p.get("maker", ""))) if p.get("maker") else ""
        aantal = len(evs)
        samenvatting = ("%d liveshow%s gepland" % (aantal, "" if aantal == 1 else "s")) if aantal else "Nog geen liveshows bekend"
        omschrijving = p["lang"] or p["kort"] or ""
        tekst = ('<p class="pbanner-tekst">%s</p>' % e(omschrijving)) if omschrijving else ""

        ld = json.dumps([jsonld_event(ev, SITE_URL.rstrip("/") + "/podcast/" + p["slug"] + ".html") for ev in evs],
                        ensure_ascii=False, indent=1)
        extra = ((leaflet_head("../") + "\n") if heeft_kaart else "") + \
                ('<script type="application/ld+json">%s</script>' % ld if evs else "")

        html_uit = (kop("%s live - Podcast Liveshows" % p["naam"],
                        (p["kort"] or ("Alle liveshows van %s." % p["naam"]))[:180],
                        basis="../", actief="catalogus", extra_head=extra,
                        pad="podcast/%s.html" % p["slug"])
                    + '  <div class="%s" style="%s">\n    %s\n'
                      '    <div class="pbanner-vak">\n      %s\n      <h1>%s</h1>\n      %s\n'
                      '      <p class="cijfers">%s &middot; %s</p>\n      %s\n    </div>\n  </div>\n'
                      % (pbanner_klasse, e(band_stijl), logo, rang, e(p["naam"]), maker, e(p["thema"]), e(samenvatting), tekst)
                    + ('  <div class="kolommen">\n    <div class="lijstkolom">\n'
                       + "\n".join(rijen) + "\n    </div>\n" + kaartblok + "  </div>\n"
                       if evs else '  <p class="leeg">Voor deze podcast staan nog geen liveshows in de agenda.</p>\n')
                    + '  <a class="terug" href="../catalogus.html">&larr; Alle podcasts</a>\n'
                    + voet(gecheckt,
                           ((leaflet_scripts("../") + "\n") if heeft_kaart else "")
                           + '<script src="../assets/favorieten.js"></script>\n'
                           + '<script>window.PAGINA = ' + json.dumps(data, ensure_ascii=False, default=str) + ';</script>\n'
                           + '<script src="../assets/podcast.js"></script>'))
        schrijf("podcast/%s.html" % p["slug"], html_uit)
        gemaakt += 1
    return gemaakt

def bouw_sitemap(podcasts):
    paden = ["index.html", "catalogus.html", "toplijst.html"] + \
            ["podcast/%s.html" % p["slug"] for p in podcasts.values()]
    vandaag = date.today().isoformat()
    regels = ['<?xml version="1.0" encoding="UTF-8"?>',
              '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for pad in paden:
        regels.append("  <url><loc>%s/%s</loc><lastmod>%s</lastmod></url>"
                      % (SITE_URL.rstrip("/"), pad, vandaag))
    regels.append("</urlset>")
    schrijf("sitemap.xml", "\n".join(regels) + "\n")
    schrijf("robots.txt", "User-agent: *\nAllow: /\n\nSitemap: %s/sitemap.xml\n" % SITE_URL.rstrip("/"))
    return len(paden)

def main():
    podcasts, shows, venues, events, status = lees_alles()

    top = lees_top100()
    if top:
        op_sleutel = {zoeksleutel(r[1]): r for r in top["lijst"]}
        for p in podcasts.values():
            tref = op_sleutel.get(zoeksleutel(p["naam"]))
            p["maker"] = tref[2] if tref else ""
    else:
        for p in podcasts.values(): p["maker"] = ""

    gecheckt = max([ev["gecheckt"] for ev in events if ev["gecheckt"]] or [""])

    bouw_index(podcasts, venues, events, gecheckt)
    bouw_catalogus(podcasts, gecheckt)
    met_live, nagekeken = bouw_toplijst(podcasts, status, gecheckt)
    n_pagina = bouw_podcastpaginas(podcasts, gecheckt)
    n_sitemap = bouw_sitemap(podcasts)

    zonder = sum(1 for v in venues.values() if not v["opkaart"])
    met_prijs = sum(1 for ev in events if ev["prijs"] is not None)
    met_tijd = sum(1 for ev in events if ev["tijd"])
    print("Site gebouwd in", HIER)
    print("  index.html           %d events" % len(events))
    print("  catalogus.html       %d podcasts" % len(podcasts))
    print("  toplijst.html        %d nagekeken, %d met liveshow, %d te doen"
          % (nagekeken, met_live, 100 - nagekeken))
    print("  podcast/             %d pagina's" % n_pagina)
    print("  sitemap.xml          %d adressen" % n_sitemap)
    print("  zalen zonder coordinaten: %d van %d" % (zonder, len(venues)))
    print("  events met aanvangstijd:  %d van %d" % (met_tijd, len(events)))
    print("  events met prijs:         %d van %d" % (met_prijs, len(events)))

    # Werklijst: alleen wat binnenkort speelt hoeft een tijd en prijs te hebben.
    # Alles verrijken schaalt niet; over een jaar is de prijs toch veranderd.
    from datetime import timedelta
    grens = (date.today() + timedelta(days=90)).isoformat()
    vandaag = date.today().isoformat()
    werk = [ev for ev in events if vandaag <= ev["iso"] <= grens
            and (not ev["tijd"] or ev["prijs"] is None)]
    if werk:
        print("\n  Nog aan te vullen (speelt binnen 90 dagen):")
        for ev in werk[:15]:
            mist = ", ".join(x for x, y in [("tijd", ev["tijd"]), ("prijs", ev["prijs"])] if not y)
            print("    %s  %-34s %-22s mist %s"
                  % (ev["iso"], ev["show"]["titel"][:34], ev["zaal"]["stad"], mist))
        if len(werk) > 15:
            print("    ... en nog %d" % (len(werk) - 15))
    else:
        print("\n  Alles wat binnen 90 dagen speelt heeft een tijd en een prijs.")

if __name__ == "__main__":
    main()
