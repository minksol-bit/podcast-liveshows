# -*- coding: utf-8 -*-
"""Leest een bronbestand in en voegt het toe aan data/podcast-liveshows.xlsx.

    python3 importeer.py data/bronnen/tonny-media.json            # laat zien wat er zou gebeuren
    python3 importeer.py data/bronnen/tonny-media.json --schrijf  # en doet het ook

Wat het script garandeert:
  - Een show die er al staat wordt niet nog een keer toegevoegd. Een event is
    hetzelfde als show, zaal en datum gelijk zijn.
  - Een zaal wordt herkend aan zijn naam of aan een van zijn aliassen, samen met
    de stad. Zo wordt "Musis" niet een tweede rij naast "Musis Sacrum".
  - Lege velden worden aangevuld, gevulde velden worden alleen overschreven als
    de bron iets anders zegt - en dat wordt dan gemeld.

Bronbestand (JSON):
{
  "bron": "Tonny Media",
  "gecheckt": "2026-08-31",
  "shows": [{
     "podcast": "Geuze & Gorgels",
     "titel": "Geuze & Gorgels Live",
     "type": "theatershow",
     "organisator": "Tonny Media",
     "bron_url": "https://...",
     "events": [
       {"datum": "2026-10-01", "tijd": "20:15", "zaal": "Musis", "stad": "Arnhem",
        "provincie": "Gelderland", "ticket_url": "https://...", "prijs_vanaf": 24.5}
     ]}]
}
"""
import json, os, re, sys, unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font

HIER = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HIER, "data", "podcast-liveshows.xlsx")
TOP100 = os.path.join(HIER, "data", "apple-top100.json")

def sleutel(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    s = s.replace("&", " en ")
    s = re.sub(r"\b(de|het|een|the|en|and|theater|schouwburg)\b", " ", s)
    return re.sub(r"[^a-z0-9]+", "", s)

def leeg(v):
    return v is None or str(v).strip() == ""

class Tabblad:
    def __init__(self, ws):
        self.ws = ws
        self.kop = [c.value for c in ws[1]]
        self.i = {k: n for n, k in enumerate(self.kop)}
        self.rijen = [r for r in ws.iter_rows(min_row=2, max_col=len(self.kop))
                      if not all(leeg(c.value) for c in r)]
    def waarde(self, rij, kolom):
        return rij[self.i[kolom]].value
    def zet(self, rij, kolom, waarde):
        rij[self.i[kolom]].value = waarde
        rij[self.i[kolom]].font = Font(name="Arial")
    def nieuw_id(self):
        ids = [self.waarde(r, "id") for r in self.rijen if not leeg(self.waarde(r, "id"))]
        return (max(int(x) for x in ids) + 1) if ids else 1
    def voegtoe(self, gegevens):
        rij_nr = self.ws.max_row + 1
        for kolom, waarde in gegevens.items():
            if kolom in self.i:
                c = self.ws.cell(row=rij_nr, column=self.i[kolom] + 1, value=waarde)
                c.font = Font(name="Arial")
        nieuwe = tuple(self.ws.cell(row=rij_nr, column=n + 1) for n in range(len(self.kop)))
        self.rijen.append(nieuwe)
        return nieuwe

def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    bronbestand = sys.argv[1]
    schrijven = "--schrijf" in sys.argv
    bron = json.load(open(bronbestand, encoding="utf-8"))
    gecheckt = bron.get("gecheckt", "")

    wb = load_workbook(XLSX)
    T = {n: Tabblad(wb[n]) for n in ["podcasts", "shows", "show_podcasts", "events", "venues"]}

    apple = {}
    if os.path.exists(TOP100):
        top = json.load(open(TOP100, encoding="utf-8"))
        apple = {sleutel(r[1]): (top["beeld_basis"], r) for r in top["lijst"]}

    # ---- zoekregisters ----
    podcast_op = {}
    for r in T["podcasts"].rijen:
        podcast_op[sleutel(T["podcasts"].waarde(r, "naam"))] = r
    show_op = {}
    for r in T["shows"].rijen:
        show_op[sleutel(T["shows"].waarde(r, "titel"))] = r
    zaal_op = {}
    for r in T["venues"].rijen:
        stad = sleutel(T["venues"].waarde(r, "stad"))
        namen = [T["venues"].waarde(r, "naam")]
        namen += [a for a in str(T["venues"].waarde(r, "aliassen") or "").split(";") if a.strip()]
        for naam in namen:
            zaal_op[(sleutel(naam), stad)] = r
    event_op = {}
    for r in T["events"].rijen:
        d = str(T["events"].waarde(r, "aanvang") or "")[:10]
        event_op[(T["events"].waarde(r, "show_id"), T["events"].waarde(r, "venue_id"), d)] = r

    verslag = {"podcast_nieuw": [], "show_nieuw": [], "zaal_nieuw": [],
               "event_nieuw": [], "event_bijgewerkt": [], "event_ongewijzigd": 0,
               "let_op": []}

    for s in bron["shows"]:
        # --- podcast ---
        pk = sleutel(s["podcast"])
        prij = podcast_op.get(pk)
        if prij is None:
            nieuw = {"id": T["podcasts"].nieuw_id(), "naam": s["podcast"]}
            if pk in apple:
                basis, a = apple[pk]
                nieuw.update({"cover_url": basis + a[5] + "/600x600bb.png",
                              "apple_id": a[4], "apple_rang": a[0]})
                verslag["let_op"].append("podcast %s staat op #%d in de Apple top 100 - thema nog invullen"
                                         % (s["podcast"], a[0]))
            else:
                verslag["let_op"].append("podcast %s: geen cover en geen thema, met de hand aanvullen"
                                         % s["podcast"])
            prij = T["podcasts"].voegtoe(nieuw)
            podcast_op[pk] = prij
            verslag["podcast_nieuw"].append(s["podcast"])
        podcast_id = T["podcasts"].waarde(prij, "id")

        # --- show ---
        sk = sleutel(s["titel"])
        srij = show_op.get(sk)
        if srij is None:
            srij = T["shows"].voegtoe({"id": T["shows"].nieuw_id(), "titel": s["titel"],
                                       "type": s.get("type", "theatershow"),
                                       "organisator": s.get("organisator", "")})
            show_op[sk] = srij
            verslag["show_nieuw"].append(s["titel"])
        show_id = T["shows"].waarde(srij, "id")

        # --- koppeling ---
        bestaat = any(T["show_podcasts"].waarde(r, "show_id") == show_id and
                      T["show_podcasts"].waarde(r, "podcast_id") == podcast_id
                      for r in T["show_podcasts"].rijen)
        if not bestaat:
            T["show_podcasts"].voegtoe({"show_id": show_id, "podcast_id": podcast_id})

        # --- events ---
        for ev in s["events"]:
            zk = (sleutel(ev["zaal"]), sleutel(ev.get("stad", "")))
            zrij = zaal_op.get(zk)
            if zrij is None:
                zrij = T["venues"].voegtoe({"id": T["venues"].nieuw_id(), "naam": ev["zaal"],
                                            "stad": ev.get("stad", ""), "provincie": ev.get("provincie", "")})
                zaal_op[zk] = zrij
                verslag["zaal_nieuw"].append("%s, %s" % (ev["zaal"], ev.get("stad", "")))
                verslag["let_op"].append("zaal %s (%s) heeft nog geen lat/lon"
                                         % (ev["zaal"], ev.get("stad", "")))
            venue_id = T["venues"].waarde(zrij, "id")

            aanvang = ev["datum"] + ((" " + ev["tijd"]) if ev.get("tijd") else "")
            sleutel_ev = (show_id, venue_id, ev["datum"])
            erij = event_op.get(sleutel_ev)
            if erij is None:
                erij = T["events"].voegtoe({
                    "id": T["events"].nieuw_id(), "show_id": show_id, "venue_id": venue_id,
                    "aanvang": aanvang, "ticket_url": ev.get("ticket_url", ""),
                    "status": ev.get("status", "in verkoop"),
                    "bron_url": ev.get("bron_url", s.get("bron_url", "")),
                    "laatst_gecheckt": gecheckt,
                    "prijs_vanaf": ev.get("prijs_vanaf"), "tijd_bron": ev.get("tijd_bron", "")})
                event_op[sleutel_ev] = erij
                verslag["event_nieuw"].append("%s - %s, %s" % (ev["datum"], ev["zaal"], s["titel"]))
            else:
                veranderd = []
                for kolom, waarde in [("aanvang", aanvang), ("ticket_url", ev.get("ticket_url")),
                                      ("prijs_vanaf", ev.get("prijs_vanaf")), ("status", ev.get("status"))]:
                    if waarde in (None, "") or kolom not in T["events"].i: continue
                    huidig = T["events"].waarde(erij, kolom)
                    if leeg(huidig):
                        T["events"].zet(erij, kolom, waarde); veranderd.append(kolom + " ingevuld")
                    elif str(huidig).strip() != str(waarde).strip():
                        veranderd.append("%s: %s -> %s (niet overschreven)" % (kolom, huidig, waarde))
                T["events"].zet(erij, "laatst_gecheckt", gecheckt)
                if veranderd:
                    verslag["event_bijgewerkt"].append("%s %s: %s" % (ev["datum"], ev["zaal"], "; ".join(veranderd)))
                else:
                    verslag["event_ongewijzigd"] += 1

    print("Bron:", bron.get("bron", bronbestand))
    for kop, sleutelnaam in [("nieuwe podcasts", "podcast_nieuw"), ("nieuwe shows", "show_nieuw"),
                             ("nieuwe zalen", "zaal_nieuw"), ("nieuwe events", "event_nieuw"),
                             ("bijgewerkte events", "event_bijgewerkt")]:
        lijst = verslag[sleutelnaam]
        print("  %-20s %d" % (kop, len(lijst)))
        for x in lijst[:40]:
            print("      -", x)
        if len(lijst) > 40:
            print("      ... en nog %d" % (len(lijst) - 40))
    print("  %-20s %d" % ("ongewijzigd", verslag["event_ongewijzigd"]))
    if verslag["let_op"]:
        print("  LET OP:")
        for x in verslag["let_op"]:
            print("      -", x)

    if schrijven:
        wb.save(XLSX)
        print("\nOpgeslagen. Draai nu: python3 bouw-site.py")
    else:
        print("\nDit was een droogloop, er is niets opgeslagen.")
        print("Voeg --schrijf toe om het echt door te voeren.")

if __name__ == "__main__":
    main()
